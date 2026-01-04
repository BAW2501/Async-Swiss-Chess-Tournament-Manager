import pandas as pd
import openpyxl
import random
from pathlib import Path
import sys

MAX_BOARDS = 30

def load_tournament_data(filepath="tournament.xlsx"):
    """Load attendees and pairings from Excel file."""
    try:
        attendees = pd.read_excel(filepath, sheet_name="Attendees")
    except Exception as e:
        print(f"ERROR: Failed to load Attendees sheet: {e}")
        sys.exit(1)
    
    try:
        pairings = pd.read_excel(filepath, sheet_name="Pairings")
        if pairings.empty:
            pairings = pd.DataFrame(columns=["Round", "Board", "White Player", "Black Player", "Results White", "Results Black"])
        else:
            pairings.columns = pairings.columns.str.strip()
            print(f"Loaded {len(pairings)} existing pairing(s)")
    except Exception as e:
        print(f"Warning: Could not load Pairings sheet ({e}), starting fresh")
        pairings = pd.DataFrame(columns=["Round", "Board", "White Player", "Black Player", "Results White", "Results Black"])
    
    return attendees, pairings

def build_player_state(attendees, pairings):
    """Build state for each player from completed games only."""
    player_state = {}
    
    for _, row in attendees.iterrows():
        player_id = int(row["ID"])
        player_state[player_id] = {
            "name": f"{row['First Name']} {row['Last Name']}",
            "points": 0.0,
            "games_played": 0,
            "last_color": None,
            "color_history": [],
            "opponents": set(),
            "opponent_scores": []
        }
    
    for _, pairing in pairings.iterrows():
        white_id = int(pairing["White Player"])
        black_id = int(pairing["Black Player"])
        result_white = pairing["Results White"]
        result_black = pairing["Results Black"]
        
        if white_id not in player_state or black_id not in player_state:
            continue
        
        if pd.notna(result_white) and pd.notna(result_black):
            player_state[white_id]["points"] += float(result_white)
            player_state[black_id]["points"] += float(result_black)
            player_state[white_id]["games_played"] += 1
            player_state[black_id]["games_played"] += 1
            player_state[white_id]["last_color"] = "White"
            player_state[black_id]["last_color"] = "Black"
            player_state[white_id]["color_history"].append("White")
            player_state[black_id]["color_history"].append("Black")
            player_state[white_id]["opponents"].add(black_id)
            player_state[black_id]["opponents"].add(white_id)
            player_state[white_id]["opponent_scores"].append(black_id)
            player_state[black_id]["opponent_scores"].append(white_id)
    
    # Calculate Buchholz scores
    for player_id in player_state:
        opponent_ids = player_state[player_id]["opponent_scores"]
        player_state[player_id]["opponent_scores"] = [
            player_state[opp_id]["points"] for opp_id in opponent_ids
        ]
    
    return player_state

def get_lowest_incomplete_round(pairings):
    """Find the lowest round number with any incomplete game."""
    if pairings.empty:
        return 1
    
    incomplete_rounds = []
    for _, pairing in pairings.iterrows():
        if pd.isna(pairing["Results White"]) or pd.isna(pairing["Results Black"]):
            incomplete_rounds.append(int(pairing["Round"]))
    
    return min(incomplete_rounds) if incomplete_rounds else (pairings["Round"].max() + 1)

def get_players_in_pending_pairings(pairings):
    """Get set of player IDs in unfinished pairings."""
    pending = set()
    for _, pairing in pairings.iterrows():
        if pd.isna(pairing["Results White"]) or pd.isna(pairing["Results Black"]):
            pending.add(int(pairing["White Player"]))
            pending.add(int(pairing["Black Player"]))
    return pending

def would_force_triple_color(player_state, player_id, new_color):
    """Check if assigning new_color would create 3 consecutive same colors."""
    history = player_state[player_id]["color_history"]
    if len(history) < 2:
        return False
    
    if history[-1] == history[-2] == new_color:
        return True
    
    return False

def can_pair(p1_state, p2_state, p2_id):
    """Check if two players can be paired (no rematches only)."""
    return p2_id not in p1_state["opponents"]

def assign_colors(p1_id, p2_id, player_state):
    """Assign colors with soft preference and hard triple-color check."""
    p1_color = player_state[p1_id]["last_color"]
    p2_color = player_state[p2_id]["last_color"]
    
    # Both new: random
    if p1_color is None and p2_color is None:
        return (p1_id, p2_id) if random.choice([True, False]) else (p2_id, p1_id)
    
    # One new, one experienced: new player gets opposite color
    if p1_color is None:
        return (p1_id, p2_id) if p2_color == "White" else (p2_id, p1_id)
    if p2_color is None:
        return (p2_id, p1_id) if p1_color == "White" else (p1_id, p2_id)
    
    # Both experienced: try to alternate
    if p1_color != p2_color:
        white_id = p1_id if p1_color == "Black" else p2_id
        black_id = p2_id if white_id == p1_id else p1_id
    else:
        # Same last color: check for triple-color violation
        if p1_color == "White":
            # Try p1=Black, p2=White
            if not would_force_triple_color(player_state, p1_id, "Black") and \
               not would_force_triple_color(player_state, p2_id, "White"):
                white_id, black_id = p2_id, p1_id
            elif not would_force_triple_color(player_state, p1_id, "White") and \
                 not would_force_triple_color(player_state, p2_id, "Black"):
                white_id, black_id = p1_id, p2_id
            else:
                # No valid assignment (hard limit)
                return None
        else:
            # p1_color == "Black"
            if not would_force_triple_color(player_state, p1_id, "White") and \
               not would_force_triple_color(player_state, p2_id, "Black"):
                white_id, black_id = p1_id, p2_id
            elif not would_force_triple_color(player_state, p1_id, "Black") and \
                 not would_force_triple_color(player_state, p2_id, "White"):
                white_id, black_id = p2_id, p1_id
            else:
                return None
    
    # Final hard check
    if would_force_triple_color(player_state, white_id, "White") or \
       would_force_triple_color(player_state, black_id, "Black"):
        return None
    
    return white_id, black_id

def generate_pairings(player_state, pending_players, rmin):
    """Generate pairings within sliding two-round window."""
    rmax = rmin + 1
    
    # Filter: only players with games_played < rmax (can play in Rmin or Rmin+1)
    available = [
        pid for pid in player_state.keys()
        if pid not in pending_players and player_state[pid]["games_played"] < rmax
    ]
    
    if len(available) < 2:
        print(f"\nOnly {len(available)} player(s) available for pairing.")
        return []
    
    print(f"\n{'='*60}")
    print(f"PAIRING: Window R{rmin}–R{rmax}, {len(available)} players available")
    print(f"{'='*60}")
    
    # Sort by points (desc)
    available.sort(key=lambda p: player_state[p]["points"], reverse=True)
    
    pairings = []
    paired = set()
    
    for i, p1 in enumerate(available):
        if p1 in paired:
            continue
        
        p1_state = player_state[p1]
        p1_games = p1_state["games_played"]
        
        # Try to find opponent with same games_played first
        for p2 in available[i+1:]:
            if p2 in paired:
                continue
            
            p2_state = player_state[p2]
            p2_games = p2_state["games_played"]
            
            if p1_games != p2_games:
                continue
            
            if can_pair(p1_state, p2_state, p2):
                colors = assign_colors(p1, p2, player_state)
                if colors is None:
                    continue
                
                white_id, black_id = colors
                round_num = p1_games + 1
                
                pairings.append({
                    "Round": round_num,
                    "Board": None,
                    "White Player": white_id,
                    "Black Player": black_id
                })
                
                paired.add(p1)
                paired.add(p2)
                break
        
        # If still unpaired, try cross-pairing
        if p1 not in paired:
            for p2 in available[i+1:]:
                if p2 in paired:
                    continue
                
                p2_state = player_state[p2]
                
                if can_pair(p1_state, p2_state, p2):
                    colors = assign_colors(p1, p2, player_state)
                    if colors is None:
                        continue
                    
                    white_id, black_id = colors
                    round_num = max(p1_games, p2_state["games_played"]) + 1
                    
                    pairings.append({
                        "Round": round_num,
                        "Board": None,
                        "White Player": white_id,
                        "Black Player": black_id
                    })
                    
                    paired.add(p1)
                    paired.add(p2)
                    break
    
    unpaired = len(available) - len(paired)
    print(f"Generated {len(pairings)} pairing(s), {unpaired} player(s) unpaired")
    
    return pairings

def assign_boards_to_pairings(pairings, existing_pairings):
    """Assign board numbers to pairings."""
    if not pairings:
        return pairings
    
    busy_boards = set()
    for _, pairing in existing_pairings.iterrows():
        if pd.isna(pairing["Results White"]) or pd.isna(pairing["Results Black"]):
            board = pairing["Board"]
            if pd.notna(board) and board != "?":
                busy_boards.add(int(board))
    
    print(f"\nBoard assignment: {len(busy_boards)} boards currently busy")
    
    if len(busy_boards) >= MAX_BOARDS:
        print(f"Tournament Full: Waiting for a board to clear.")
        for pairing in pairings:
            pairing["Board"] = "?"
        return pairings
    
    for pairing in pairings:
        assigned = False
        for board_num in range(1, MAX_BOARDS + 1):
            if board_num not in busy_boards:
                pairing["Board"] = board_num
                busy_boards.add(board_num)
                assigned = True
                break
        
        if not assigned:
            pairing["Board"] = "?"
    
    return pairings

def append_pairings_to_excel(filepath, new_pairings, player_state):
    """Append new pairings to Excel."""
    if not new_pairings:
        print("\nNo new pairings to append")
        return
    
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Pairings"]
    
    for pairing in new_pairings:
        white_id = pairing["White Player"]
        black_id = pairing["Black Player"]
        
        ws.append([
            pairing["Round"],
            pairing["Board"],
            white_id,
            player_state[white_id]["name"],
            black_id,
            player_state[black_id]["name"],
            None,
            None
        ])
    
    wb.save(filepath)
    print(f"✓ Appended {len(new_pairings)} pairing(s)")

def calculate_standings(player_state):
    """Calculate standings with Buchholz tie-break."""
    standings = []
    
    for player_id, state in player_state.items():
        buchholz = sum(state["opponent_scores"])
        
        standings.append({
            "Pos": 0,
            "Player Name": state["name"],
            "Pt": state["points"],
            "BucT": buchholz
        })
    
    standings.sort(key=lambda x: (x["Pt"], x["BucT"]), reverse=True)
    
    for i, standing in enumerate(standings):
        standing["Pos"] = i + 1
    
    return standings

def write_standings_to_excel(filepath, standings):
    """Write standings to Excel."""
    wb = openpyxl.load_workbook(filepath)
    
    if "Standings" in wb.sheetnames:
        del wb["Standings"]
    
    ws = wb.create_sheet("Standings")
    ws.append(["Pos", "Player Name", "Pt", "BucT"])
    
    for standing in standings:
        ws.append([
            standing["Pos"],
            standing["Player Name"],
            standing["Pt"],
            standing["BucT"]
        ])
    
    wb.save(filepath)
    print("✓ Updated standings")

def main():
    filepath = "tournament.xlsx"
    
    if not Path(filepath).exists():
        print(f"ERROR: {filepath} not found!")
        sys.exit(1)
    
    print("=" * 60)
    print("Swiss-Style Overlapping Round Tournament System")
    print("=" * 60)
    
    attendees, pairings = load_tournament_data(filepath)
    player_state = build_player_state(attendees, pairings)
    
    rmin = get_lowest_incomplete_round(pairings)
    print(f"\nLowest incomplete round: R{rmin}")
    
    pending_players = get_players_in_pending_pairings(pairings)
    print(f"{len(pending_players)} player(s) currently in unfinished games")
    
    new_pairings = generate_pairings(player_state, pending_players, rmin)
    
    if new_pairings:
        new_pairings = assign_boards_to_pairings(new_pairings, pairings)
        append_pairings_to_excel(filepath, new_pairings, player_state)
    else:
        print("\nNo new pairings generated")
    
    standings = calculate_standings(player_state)
    write_standings_to_excel(filepath, standings)
    
    print("\n" + "=" * 60)
    print("✓ SUCCESS!")
    print("=" * 60)

if __name__ == "__main__":
    main()